#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Extracteur de données réelles pour le seed gestRH.

Lit `SIRH_UNIVERSITE_ST_CHRISTOPHER_.xlsx` (à la racine du projet, NON commité —
PII) et produit `prisma/seed-data.json` (NON commité non plus). Le seed.ts lit
ce JSON pour peupler la base. Ce script, lui, ne contient aucune donnée
personnelle et peut donc être commité pour régénérer le seed.

Usage :  python scripts/seed-extract.py
"""
import json
import re
import unicodedata
from datetime import datetime, date

import openpyxl

SRC = "SIRH_UNIVERSITE_ST_CHRISTOPHER_.xlsx"
OUT = "prisma/seed-data.json"
DOMAIN = "stchris.edu"

DG_SERVICE = "Direction Générale"
PER_SERVICE = "Corps Enseignant (PER)"
PREST_SERVICE = "Prestataires Académiques"
LEADERSHIP_SERVICE = "Rectorat & Doyenné"


# ---------------------------------------------------------------------------
def iso(v):
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return None


def clean(v):
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def strip_accents(s):
    return "".join(
        c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn"
    )


def norm_token_set(s):
    """Ensemble de tokens normalisés d'un nom (sans accents, civilité, **)."""
    s = strip_accents(clean(s)).lower()
    s = s.replace("*", " ")
    s = re.sub(r"\b(mme|mr|m|mlle|dr|pr)\b", " ", s)
    toks = re.findall(r"[a-z]+", s)
    return set(t for t in toks if len(t) > 1)


def gender(v):
    g = clean(v).upper()
    return "FEMME" if g.startswith("F") else "HOMME"


def contract_type(v, default="CDI"):
    t = clean(v).upper()
    if "PRESTAT" in t:
        return "PRESTATION"
    if "STAGE" in t:
        return "STAGE"
    if "VACAT" in t:
        return "VACATAIRE"
    if "CDD" in t:
        return "CDD"
    if "CDI" in t:
        return "CDI"
    return default


def to_float(v):
    try:
        if v is None or clean(v) == "":
            return None
        return float(v)
    except (TypeError, ValueError):
        return None


def email_from(first, last, used):
    base = strip_accents(f"{first}.{last}").lower()
    base = re.sub(r"[^a-z.]+", "", base).strip(".") or "agent"
    e = f"{base}@{DOMAIN}"
    i = 2
    while e in used:
        e = f"{base}{i}@{DOMAIN}"
        i += 1
    used.add(e)
    return e


def uniq_mat(mat):
    """Matricule unique (la source contient quelques doublons)."""
    if mat not in used_matricules:
        used_matricules.add(mat)
        return mat
    i = 2
    while f"{mat}-{i}" in used_matricules:
        i += 1
    m = f"{mat}-{i}"
    used_matricules.add(m)
    return m


def slug_code(name, used):
    words = [w for w in re.findall(r"[A-Za-z]+", strip_accents(name))
             if w.lower() != "service"]
    s = (words[0][:5].upper() if words else "SVC")
    c = s
    i = 2
    while c in used:
        c = f"{s}{i}"
        i += 1
    used.add(c)
    return c


# ---------------------------------------------------------------------------
wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)

used_emails = set()
used_matricules = set()
agents = []  # dicts
balances_by_name = []  # (token_set, balance dict)

# --- CONGES PATS : soldes par nom ---
ws = wb["CONGES PATS"]
for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
    name = clean(row[1]) if len(row) > 1 else ""
    if not name or "nom" in name.lower():
        continue
    total = to_float(row[6]) if len(row) > 6 else None  # Total 2026
    taken = to_float(row[5]) if len(row) > 5 else None  # Congés pris 2026
    if total is None and taken is None:
        continue
    balances_by_name.append(
        (
            norm_token_set(name),
            {
                "year": 2026,
                "type": "ANNUEL",
                "totalDays": total if total is not None else 0.0,
                "usedDays": taken if taken is not None else 0.0,
            },
        )
    )


def find_balance(first, last):
    target = norm_token_set(f"{first} {last}")
    best = None
    best_score = 0
    for toks, bal in balances_by_name:
        inter = len(target & toks)
        if inter > best_score and inter >= 2:
            best_score = inter
            best = bal
    return best


# --- Effectif PATS ---
ws = wb["Effectif PATS"]
for row in ws.iter_rows(min_row=2, values_only=True):
    mat = clean(row[1]) if len(row) > 1 else ""
    last = clean(row[2]) if len(row) > 2 else ""
    first = clean(row[3]) if len(row) > 3 else ""
    if not (mat and last):
        continue
    fonction = clean(row[10]) if len(row) > 10 else ""
    service = clean(row[9]) if len(row) > 9 else "Service Administratif"
    sub = "PATS_TECHNIQUE" if re.search(r"techni", fonction, re.I) else "PATS_ADMINISTRATIF"
    email = clean(row[11]) if len(row) > 11 and clean(row[11]) else email_from(first, last, used_emails)
    email = email.lower()
    if email in used_emails and email:
        # garde l'email réel mais évite les doublons
        email = email_from(first, last, used_emails)
    else:
        used_emails.add(email)
    mat = uniq_mat(mat)
    note = to_float(row[19]) if len(row) > 19 else None
    agents.append({
        "matricule": mat, "firstName": first, "lastName": last,
        "gender": gender(row[4] if len(row) > 4 else None),
        "birthDate": iso(row[5]) if len(row) > 5 else None,
        "category": "PATS", "subCategory": sub,
        "jobTitle": fonction or "Agent", "serviceName": service,
        "email": email, "phone": clean(row[13]) if len(row) > 13 else None,
        "hireDate": iso(row[7]) if len(row) > 7 else None,
        "contract": {
            "type": contract_type(row[6] if len(row) > 6 else None, "CDI"),
            "startDate": iso(row[7]) if len(row) > 7 else None,
            "endDate": iso(row[8]) if len(row) > 8 else None,
        },
        "note": note,
        "leaveBalance": find_balance(first, last),
        "fonction": fonction,
    })

# --- PER ---
ws = wb["PER"]
for row in ws.iter_rows(min_row=6, values_only=True):
    mat = clean(row[0]) if len(row) > 0 else ""
    last = clean(row[1]) if len(row) > 1 else ""
    first = clean(row[2]) if len(row) > 2 else ""
    if not (mat and last):
        continue
    poste = clean(row[9]) if len(row) > 9 else "Enseignant"
    spec = clean(row[8]) if len(row) > 8 else ""
    email = (clean(row[18]).lower() if len(row) > 18 and clean(row[18]) else email_from(first, last, used_emails))
    if email and email not in used_emails:
        used_emails.add(email)
    mat = uniq_mat(mat)
    note = to_float(row[19]) if len(row) > 19 else None
    agents.append({
        "matricule": mat, "firstName": first, "lastName": last,
        "gender": gender(row[5] if len(row) > 5 else None),
        "birthDate": iso(row[3]) if len(row) > 3 else None,
        "category": "PER", "subCategory": "PER_ENSEIGNEMENT",
        "jobTitle": (f"{poste} — {spec}" if spec else poste),
        "serviceName": PER_SERVICE,
        "email": email, "phone": None,
        "hireDate": iso(row[13]) if len(row) > 13 else None,
        "contract": {
            "type": contract_type(row[10] if len(row) > 10 else None, "CDD"),
            "startDate": iso(row[13]) if len(row) > 13 else None,
            "endDate": iso(row[14]) if len(row) > 14 else None,
        },
        "note": note,
        "leaveBalance": None,
        "fonction": poste,
    })

# --- PRESTATAIRES ---
ws = wb["PRESTATAIRES"]
prest_n = 0
for row in ws.iter_rows(min_row=6, values_only=True):
    last = clean(row[0]) if len(row) > 0 else ""
    first = clean(row[1]) if len(row) > 1 else ""
    fonction = clean(row[5]) if len(row) > 5 else ""
    if not last or not fonction:
        continue
    prest_n += 1
    mat = f"PREST-{prest_n:04d}"
    used_matricules.add(mat)
    is_leader = bool(re.search(r"recteur|doyen", fonction, re.I))
    agents.append({
        "matricule": mat, "firstName": first or "—", "lastName": last,
        "gender": gender(row[4] if len(row) > 4 else None),
        "birthDate": iso(row[2]) if len(row) > 2 else None,
        "category": "PRESTATAIRE", "subCategory": "PRESTATAIRE_SERVICE",
        "jobTitle": fonction,
        "serviceName": LEADERSHIP_SERVICE if is_leader else PREST_SERVICE,
        "email": email_from(first or last, last, used_emails),
        "phone": None,
        "hireDate": None,
        "contract": {"type": "PRESTATION", "startDate": None, "endDate": None},
        "note": None,
        "leaveBalance": None,
        "fonction": fonction,
    })


# ---------------------------------------------------------------------------
# Services + chefs de service
# ---------------------------------------------------------------------------
service_names = []
for a in agents:
    if a["serviceName"] and a["serviceName"] not in service_names:
        service_names.append(a["serviceName"])
for forced in [DG_SERVICE, PER_SERVICE, PREST_SERVICE, LEADERSHIP_SERVICE]:
    if forced not in service_names:
        service_names.append(forced)

used_codes = set()
services = []
# Chef de service = fonction « Chef … » ou « Directeur/Directrice … »
# (on exclut volontairement « Responsable » qui désigne ici des rôles
# pédagogiques/prestataires, pas des chefs de service RH).
CHEF_RE = re.compile(r"\bchef\b|directeur|directrice", re.I)
for sname in service_names:
    members = [a for a in agents if a["serviceName"] == sname]
    manager = None
    for a in members:
        if CHEF_RE.search(a["fonction"]):
            manager = a["matricule"]
            break
    services.append({
        "name": sname,
        "code": slug_code(sname, used_codes),
        "managerMatricule": manager,
    })


def find_by_fonction(pattern):
    rx = re.compile(pattern, re.I)
    for a in agents:
        if rx.search(a["fonction"]):
            return a
    return None


dg = find_by_fonction(r"directeur g[eé]n")
recteur = find_by_fonction(r"\brecteur\b")
doyen = find_by_fonction(r"doyen ex[eé]cutif") or find_by_fonction(r"doyen")
chef = find_by_fonction(r"chef service scolarit")
# un agent « lambda » pour la démo : 1er PATS qui n'est ni chef ni direction
agent_demo = next(
    (a for a in agents
     if a["category"] == "PATS" and not CHEF_RE.search(a["fonction"])),
    None,
)

demo_users = [
    {"email": f"direction@{DOMAIN}", "role": "DIRECTION", "linkMatricule": dg["matricule"] if dg else None},
    {"email": f"recteur@{DOMAIN}", "role": "RECTEUR", "linkMatricule": recteur["matricule"] if recteur else None},
    {"email": f"doyen@{DOMAIN}", "role": "DOYEN", "linkMatricule": doyen["matricule"] if doyen else None},
    {"email": f"rh@{DOMAIN}", "role": "DRH", "linkMatricule": None},
    {"email": f"chef@{DOMAIN}", "role": "MANAGER", "linkMatricule": chef["matricule"] if chef else None},
    {"email": f"agent@{DOMAIN}", "role": "AGENT", "linkMatricule": agent_demo["matricule"] if agent_demo else None},
]

# Nettoyage du champ interne 'fonction'
for a in agents:
    a.pop("fonction", None)

data = {
    "generatedFrom": SRC,
    "services": services,
    "agents": agents,
    "demoUsers": demo_users,
    "demoPassword": "sirh2026",
}

with open(OUT, "w", encoding="utf-8") as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

# Résumé (sans PII)
by_cat = {}
for a in agents:
    by_cat[a["category"]] = by_cat.get(a["category"], 0) + 1
with_balance = sum(1 for a in agents if a["leaveBalance"])
with_note = sum(1 for a in agents if a["note"] is not None)
print(f"OK -> {OUT}")
print(f"  services        : {len(services)}")
print(f"  agents          : {len(agents)}  {by_cat}")
print(f"  soldes congés    : {with_balance}")
print(f"  notes éval       : {with_note}")
print(f"  chefs de service : {sum(1 for s in services if s['managerMatricule'])}")
print(f"  demo: DG={'ok' if dg else 'X'} Recteur={'ok' if recteur else 'X'} "
      f"Doyen={'ok' if doyen else 'X'} Chef={'ok' if chef else 'X'} Agent={'ok' if agent_demo else 'X'}")
